"""
Job Hub — FastAPI backend (deploy-anywhere).

Endpoints
  GET  /api/health        liveness + which model / db is configured
  GET  /api/data/{key}    read a stored JSON blob
  PUT  /api/data/{key}    write a stored JSON blob
  GET  /api/export.xlsx   download all applications as Excel
  POST /api/claude        proxy to Anthropic (API key stays server-side)

Storage is a single key-value table behind SQLAlchemy, so the SAME code runs on
SQLite locally and Postgres in production — switch with one env var:
  DATABASE_URL=sqlite:///./jobhub.db                 (default, local)
  DATABASE_URL=postgresql://user:pass@host:5432/db   (hosted, accessible anywhere)

In production the backend also serves the built frontend (set STATIC_DIR), so the
whole app is one service at one URL — open it from any device.
"""
import os
import io
import json
import uuid
import httpx
from datetime import datetime, timezone
from fastapi import FastAPI, Request, Response, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from sqlalchemy import create_engine, Table, Column, String, Text, Integer, LargeBinary, MetaData, select, insert, update, delete
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

load_dotenv()

API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()
MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6").strip()
MAX_TOKENS_CAP = int(os.environ.get("MAX_TOKENS_CAP", "1500"))
ALLOWED_ORIGINS = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if o.strip()]
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./jobhub.db")
# Some hosts (Render, Heroku) give a "postgres://" URL; SQLAlchemy needs "postgresql://".
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
STATIC_DIR = os.environ.get("STATIC_DIR", "").strip()
# Optional shared password. If set, every /api call must send it (header
# X-App-Password). Leave empty to run with no login (e.g. local only).
APP_PASSWORD = os.environ.get("APP_PASSWORD", "").strip()

ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"

# ---- Database (SQLite or Postgres, same code) ------------------------------
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, future=True, pool_pre_ping=True, connect_args=connect_args)
metadata = MetaData()
store = Table("store", metadata, Column("key", String(255), primary_key=True), Column("value", Text))
files = Table(
    "files", metadata,
    Column("id", String(64), primary_key=True),
    Column("name", String(512)),
    Column("content_type", String(255)),
    Column("size", Integer),
    Column("text", Text),            # extracted plain text (powers the AI tools)
    Column("data", LargeBinary),     # the original file bytes
    Column("created_at", String(40)),
)
metadata.create_all(engine)


def get_value(key: str):
    with engine.begin() as conn:
        row = conn.execute(select(store.c.value).where(store.c.key == key)).fetchone()
        return row[0] if row else None


def set_value(key: str, value_str: str):
    with engine.begin() as conn:
        exists = conn.execute(select(store.c.key).where(store.c.key == key)).fetchone()
        if exists:
            conn.execute(update(store).where(store.c.key == key).values(value=value_str))
        else:
            conn.execute(insert(store).values(key=key, value=value_str))


# ---- App -------------------------------------------------------------------
app = FastAPI(title="Job Hub backend")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Paths under /api that never require the password (liveness + the login probe).
OPEN_API_PATHS = {"/api/health"}


@app.middleware("http")
async def require_password(request: Request, call_next):
    """When APP_PASSWORD is set, gate every /api call (except OPEN_API_PATHS and
    CORS preflights). The frontend sends the password in the X-App-Password header."""
    path = request.url.path
    if (APP_PASSWORD and request.method != "OPTIONS"
            and path.startswith("/api/") and path not in OPEN_API_PATHS):
        if request.headers.get("x-app-password", "") != APP_PASSWORD:
            return JSONResponse({"detail": "Unauthorized"}, status_code=401)
    return await call_next(request)


@app.get("/api/health")
def health():
    return {"ok": True, "model": MODEL, "key_loaded": bool(API_KEY), "database": DATABASE_URL.split("://")[0], "auth": bool(APP_PASSWORD)}


@app.get("/api/login")
def login():
    # Reaching here means the middleware accepted the password (or none is set).
    return {"ok": True}


@app.get("/api/data/{key}")
def get_data(key: str):
    v = get_value(key)
    return {"value": json.loads(v) if v else None}


@app.put("/api/data/{key}")
async def put_data(key: str, request: Request):
    payload = await request.json()
    set_value(key, json.dumps(payload.get("value")))
    return {"ok": True}


# ---- Resume files: extract text + store/list/download/delete ---------------
def extract_text(name: str, data: bytes) -> str:
    """Pull plain text out of a PDF / DOCX / TXT. Raises HTTPException(400) for
    unsupported types; lets parsing errors bubble up to the caller."""
    name = (name or "").lower()
    if name.endswith(".pdf"):
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        return "\n".join((page.extract_text() or "") for page in reader.pages).strip()
    if name.endswith(".docx"):
        import docx
        document = docx.Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs).strip()
    if name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="ignore").strip()
    raise HTTPException(status_code=400, detail="Unsupported file type. Upload a PDF, DOCX, or TXT.")


@app.post("/api/extract")
async def extract(file: UploadFile = File(...)):
    """Extract text only (no storage) — used by the quick-upload on the compare tabs."""
    data = await file.read()
    if len(data) > 8 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (8 MB max).")
    try:
        text = extract_text(file.filename, data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not read the file: {e}")
    if not text:
        raise HTTPException(status_code=422, detail="No text found. If this is a scanned/image PDF, paste the text manually.")
    return {"text": text, "filename": file.filename}


@app.post("/api/files")
async def upload_file(file: UploadFile = File(...)):
    """Store a resume file (bytes + extracted text) in the database."""
    data = await file.read()
    if len(data) > 8 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (8 MB max).")
    try:
        text = extract_text(file.filename, data)
    except HTTPException:
        raise
    except Exception:
        text = ""  # keep the file even if its text can't be read (e.g. scanned PDF)
    fid = uuid.uuid4().hex
    created = datetime.now(timezone.utc).isoformat()
    with engine.begin() as conn:
        conn.execute(insert(files).values(
            id=fid, name=file.filename or "resume",
            content_type=file.content_type or "application/octet-stream",
            size=len(data), text=text, data=data, created_at=created,
        ))
    return {"id": fid, "name": file.filename, "size": len(data), "created_at": created, "hasText": bool(text)}


@app.get("/api/files")
def list_files():
    with engine.begin() as conn:
        rows = conn.execute(
            select(files.c.id, files.c.name, files.c.size, files.c.created_at, files.c.text)
            .order_by(files.c.created_at.desc())
        ).fetchall()
    return {"files": [
        {"id": r[0], "name": r[1], "size": r[2], "created_at": r[3], "hasText": bool(r[4] and r[4].strip())}
        for r in rows
    ]}


@app.get("/api/files/{fid}/text")
def file_text(fid: str):
    with engine.begin() as conn:
        row = conn.execute(select(files.c.text).where(files.c.id == fid)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="File not found.")
    return {"text": row[0] or ""}


@app.get("/api/files/{fid}")
def download_file(fid: str):
    with engine.begin() as conn:
        row = conn.execute(select(files.c.name, files.c.content_type, files.c.data).where(files.c.id == fid)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="File not found.")
    name, ctype, data = row
    # "inline" so Preview renders the file in the browser tab (PDFs especially)
    # instead of forcing a download.
    return StreamingResponse(
        io.BytesIO(data),
        media_type=ctype or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{name}"'},
    )


@app.delete("/api/files/{fid}")
def delete_file(fid: str):
    with engine.begin() as conn:
        conn.execute(delete(files).where(files.c.id == fid))
    return {"ok": True}


# ---- Excel export ----------------------------------------------------------
STAGE_LABELS = {"saved": "Saved", "applied": "Applied", "interview": "Interview", "offer": "Offer"}
COLUMNS = [
    "Company", "Role", "Status", "Platform", "Sponsorship", "Priority", "Role level",
    "Work model", "Location", "Comp range", "Resume version", "Date applied",
    "Follow-up", "Next step", "Contacts", "Notes", "Link", "JD match %",
]


def build_apps_workbook():
    """Build the Applications spreadsheet from the current database state."""
    raw = get_value("jah:applications")
    apps = json.loads(raw) if raw else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(COLUMNS)

    for a in apps:
        contacts = "; ".join(
            f"{c.get('name', '')} ({c.get('type', '')}" + (f", {c.get('handle')}" if c.get("handle") else "") + ")"
            for c in (a.get("contacts") or [])
        )
        analysis = a.get("analysis") or {}
        ws.append([
            a.get("company", ""), a.get("role", ""),
            STAGE_LABELS.get(a.get("stage", ""), a.get("stage", "")),
            a.get("platform", ""), a.get("sponsorship", ""), a.get("priority", ""), a.get("level", ""),
            a.get("workModel", ""), a.get("location", ""), a.get("comp", ""), a.get("resumeVersion", ""),
            a.get("dateApplied", ""), a.get("followUp", ""), a.get("nextStep", ""), contacts,
            a.get("notes", ""), a.get("link", ""),
            analysis.get("coverageScore", ""),
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, h in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(42, len(h) + 6))
    ws.freeze_panes = "A2"
    return wb


@app.get("/api/export.xlsx")
def export_xlsx():
    bio = io.BytesIO()
    build_apps_workbook().save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=job-applications.xlsx"},
    )


# ---- Anthropic proxy -------------------------------------------------------
@app.post("/api/claude")
async def claude(request: Request):
    if not API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not set on the server.")
    body = await request.json()
    body["model"] = MODEL
    body["max_tokens"] = min(int(body.get("max_tokens", 1024)), MAX_TOKENS_CAP)
    headers = {
        "x-api-key": API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            r = await client.post(ANTHROPIC_URL, json=body, headers=headers)
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Upstream request failed: {e}")
    return Response(content=r.content, status_code=r.status_code, media_type="application/json")


# ---- Serve the built frontend (production single-service) ------------------
# Registered LAST so it never shadows the /api routes above.
if STATIC_DIR and os.path.isdir(STATIC_DIR):
    @app.get("/{full_path:path}")
    def spa(full_path: str):
        candidate = os.path.join(STATIC_DIR, full_path)
        if full_path and os.path.isfile(candidate):
            return FileResponse(candidate)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))
else:
    @app.get("/")
    def root():
        return {"ok": True, "note": "API only. Set STATIC_DIR to also serve the frontend, or run the Vite dev server."}
